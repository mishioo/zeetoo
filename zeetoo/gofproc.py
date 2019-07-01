import openpyxl as opxl
import re
import argparse
from itertools import chain
from pathlib import Path
import logging as lgg

number_group = r'\s*(-?\d+\.?\d*)'
number = number_group.replace('(', '').replace(')', '')
energies = re.compile(
    r' Zero-point correction=\s*(-?\d+\.?\d*) \(Hartree/Particle\)\n'
    r' Thermal correction to Energy=\s*(-?\d+\.?\d*)\n'
    r' Thermal correction to Enthalpy=\s*(-?\d+\.?\d*)\n'
    r' Thermal correction to Gibbs Free Energy=\s*(-?\d+\.?\d*)\n'
    r' Sum of electronic and zero-point Energies=\s*(-?\d+\.?\d*)\n'
    r' Sum of electronic and thermal Energies=\s*(-?\d+\.?\d*)\n'
    r' Sum of electronic and thermal Enthalpies=\s*(-?\d+\.?\d*)\n'
    r' Sum of electronic and thermal Free Energies=\s*(-?\d+\.?\d*)'
)  # use .search(text).groups()
frequencies = re.compile(
    'Frequencies(?:\s+(?:Fr= \d+)?--\s+)' + 3 * number_group
)


def get_args(argv=None):
    prs = argparse.ArgumentParser(
        description='Process gaussian optimization files to get energies and '
                    'number of imaginary frequencies.'
    )
    prs.add_argument(
        'files', type=Path, nargs='+',
        help='One or more gaussian output files or directories with such.'
    )
    prs.add_argument(
        '-f', '--file', type=Path, default=None,
        help='Wries output to specified excel file.'
    )
    prs.add_argument(
        '-c', '--corrections', action='store_true',
        help='Include corrections in the output.'
    )
    prs.add_argument(
        '-i', '--imag_count', action='store_true',
        help='Include number of imaginary numbers in the output.'
    )
    prs.add_argument(
        '-e', '--energies', action='store_true',
        help='Include energies values in the output.'
    )
    prs.add_argument(
        '-a', '--append_entry', metavar='COL', type=int,
        help='Append output values to entry, if it already exists, otherwise '
             'add a new entry. Values will be written to sheet starting at '
             'column COL.'
    )
    log = prs.add_mutually_exclusive_group()
    log.add_argument(
        '-u', '--unconverged', action='store_true',
        help='Print names of files that did not converged.'
    )
    log.add_argument(
        '-s', '--silent', action='store_true',
        help='Suppress printing to sdtout.'
    )
    args = prs.parse_args(argv)
    if args.append_entry is not None and args.file is None:
        prs.error("--append_entry needs --file to be specified.")
    return args


def get_data(path):
    with path.open('r') as file:
        text = file.read()
    ens = energies.search(text)
    freqs = frequencies.findall(text)
    if not ens or not freqs:
        lgg.debug(f'NOT CONVERGED: {path.name}')
        return []
    else:
        imag = sum(float(f) < 0 for line in freqs for f in line)
        return [*ens.groups(), imag]


def select_data(line, args):
    entries = []
    if not line:
        return entries
    zcr, tcr, ecr, gcr, zpe, ten, ent, gib, imag = line
    # entry: name, value
    if args.energies:
        entries.append(("ZPE", float(zpe)))
        entries.append(("TEN", float(ten)))
        entries.append(("ENT", float(ent)))
        entries.append(("GIB", float(gib)))
    if args.corrections:
        entries.append(("ZPECORR", float(zcr)))
        entries.append(("TENCORR", float(tcr)))
        entries.append(("ENTCORR", float(ecr)))
        entries.append(("GIBCORR", float(gcr)))
    if args.imag_count:
        entries.append(("Imag.Freqs", int(imag)))
    return entries


def main():
    args = get_args()
    lgg.basicConfig(level='WARNING' if args.silent else 'INFO')
    dirs = (path for path in args.files if path.is_dir())
    inner_files = (
        path for dir in dirs for path in dir.iterdir()
        if path.is_file() and path.suffix in ['.log', '.out']
    )
    args_files = (
        path for path in args.files
        if path.is_file() and path.suffix in ['.log', '.out']
    )
    files = chain(args_files, inner_files)
    lines = ((path.name, get_data(path)) for path in files)
    data = ((file, select_data(line, args)) for file, line in lines)
    if args.file:
        wb = opxl.load_workbook(str(args.file))
        sheet = wb.active
        owned = {row[0].value: row[0].row for row in sheet.iter_rows(max_col=1)}
    else:
        wb, sheet, owned = None, None, {}
    unconverged = []
    length = 0
    for file, entries in data:
        if not entries:  # no data extracted
            unconverged.append(file)
            continue
        length = len(file) if len(file) > length else length
        lgg.info(
            ' '.join((
                f"{file: <{length}} -",
                *("{}= {}".format(*e) for e in entries)
            ))
        )
        if args.file and args.append_entry is not None:
            if file in owned:
                row = owned[file]
            else:
                row = sheet.max_row + 1
                sheet.cell(row=row, column=1, value=file)
            col = args.append_entry
            for n, value in enumerate(entries):
                sheet.cell(row=row, column=col+n, value=value[1])
        elif args.file:
            sheet.append([
                file, '',  # place for comment on this file
                *(e[1] for e in entries)
            ])
    if args.unconverged:
        lgg.info(f"{len(unconverged)} unconverged files:")
        for filename in unconverged:
            lgg.info(f'\t{filename}')
    if args.file:
        wb.save(str(args.file))


if __name__ == '__main__':

    main()
